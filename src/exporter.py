"""
Phase 4: Excel Exporter v3 — 17-column layout, native data types.

FIX 1: Dates as datetime, values as float/int, quantities as int
FIX 2: Freeze panes on B5
FIX 5: Value 0.01 treated as None (API placeholder)

Output: data/export/YYMMDD_TED_Tender Data_00.XX.xlsx
"""

import json
import logging
import re
import shutil
from pathlib import Path
from datetime import datetime, date
from typing import Optional, Union, Set

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Fixed FX rates to EUR ──
FX_RATES_TO_EUR = {
    "EUR": 1.0, "DKK": 0.134, "SEK": 0.087, "PLN": 0.233,
    "CZK": 0.040, "RON": 0.201, "NOK": 0.085, "GBP": 1.17,
    "CHF": 1.06, "HRK": 0.133, "BGN": 0.511, "HUF": 0.0025,
}

# ── Country normalization ──
COUNTRY_NORMALIZE = {
    "DEU": "Germany", "FRA": "France", "POL": "Poland", "ROU": "Romania",
    "CZE": "Czech Republic", "DNK": "Denmark", "SWE": "Sweden",
    "NLD": "Netherlands", "IRL": "Ireland", "BEL": "Belgium",
    "ESP": "Spain", "ITA": "Italy", "AUT": "Austria", "CHE": "Switzerland",
    "LUX": "Luxembourg", "SVN": "Slovenia", "NOR": "Norway",
    "MKD": "North Macedonia", "SVK": "Slovakia", "GBR": "United Kingdom",
    "FIN": "Finland", "HRV": "Croatia", "LTU": "Lithuania",
    "EST": "Estonia", "BGR": "Bulgaria", "HUN": "Hungary",
    "PRT": "Portugal", "GRC": "Greece", "LVA": "Latvia",
    "MLT": "Malta", "CYP": "Cyprus", "Czechia": "Czech Republic",
}


def normalize_country(raw: str) -> str:
    if not raw:
        return "Unknown"
    raw = raw.strip()
    return COUNTRY_NORMALIZE.get(raw, raw)


def clean_value(val) -> Optional[float]:
    """Convert value to float. Treat 0.01 and 0 as None (API placeholders)."""
    if val is None or val == "" or val == "nan":
        return None
    try:
        num = float(str(val).replace(",", "").replace(" ", ""))
        if num <= 0.01:
            return None
        return num
    except (ValueError, TypeError):
        return None


def clean_int(val) -> Optional[int]:
    """Convert value to int, return None if empty/invalid."""
    if val is None or val == "" or val == "null" or val == "None":
        return None
    try:
        return int(float(str(val).replace(",", "").replace(" ", "")))
    except (ValueError, TypeError):
        return None


def parse_date(date_str: str) -> Optional[date]:
    """Parse date string to datetime.date. Returns None on failure."""
    if not date_str or date_str == "None":
        return None
    # Strip timezone suffixes
    for suffix in ["+01:00", "+02:00", "+00:00", "+03:00", "Z"]:
        date_str = date_str.replace(suffix, "")
    date_str = date_str.strip()
    for fmt in ["%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y%m%d"]:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def clean_winner(winner_str) -> str:
    """Remove duplicate winner names that appear on separate lines."""
    if not winner_str or not isinstance(winner_str, str):
        return winner_str or ""
    lines = [line.strip() for line in winner_str.strip().split("\n") if line.strip()]
    unique = list(dict.fromkeys(lines))  # preserves order, removes exact dupes
    return "\n".join(unique) if len(unique) > 1 else unique[0] if unique else ""


def determine_status(notice: dict, flat: dict) -> str:
    """Determine tender status: Open / Awarded / Closed / Unknown.

    Priority order:
      1. Winner present           → Awarded
      2. Award-notice title hints → Awarded
      3. Deadline field (exact)   → Open / Closed
      4. Pub-date heuristic       → Open (<6 months) / Closed (≥6 months)
      5. Fallback                 → Unknown (should be rare after step 4)
    """
    # 1. Cleaned winner present → Awarded
    winner = flat.get("_winner_name")
    if winner and str(winner).strip() and str(winner).strip().lower() not in ("nan", "none", ""):
        return "Awarded"

    # 2. Title / notice-type hints at award notice
    title = str(flat.get("_title_final", "") or "").lower()
    raw_title = str(notice.get("title", "") or "").lower()
    combined = f"{title} {raw_title}"
    if any(x in combined for x in ["award notice", "contract award", "vergabebekanntmachung",
                                    "zuschlag", "résultat", "bekanntmachung vergebener",
                                    "- result", "attribution", "vergabe"]):
        return "Awarded"
    # Also check notice_type in raw data
    raw = notice.get("_raw") or {}
    notice_type = raw.get("notice-type", "")
    if isinstance(notice_type, dict):
        vals = list(notice_type.values())
        notice_type = str(vals[0]) if vals else ""
    if any(x in str(notice_type).lower() for x in ["award", "result", "vergabe", "résultat"]):
        return "Awarded"

    # 3. Deadline vs. today — check both normalised field and raw OCDS/TED payload
    deadline_raw = (
        notice.get("submission_deadline")
        or raw.get("deadline-receipt-tender-date-lot")
    )
    if deadline_raw:
        # Unwrap dict/list wrappers (TED v3 often nests values)
        if isinstance(deadline_raw, dict):
            vals = list(deadline_raw.values())
            deadline_raw = vals[0] if vals else None
        if isinstance(deadline_raw, list):
            deadline_raw = deadline_raw[0] if deadline_raw else None
        if deadline_raw:
            deadline_date = parse_date(str(deadline_raw)[:10])
            if deadline_date:
                return "Open" if deadline_date >= date.today() else "Closed"

    # 4. Publication-date heuristic (eliminates most "Unknown")
    #    - < 6 months old, no winner → probably still open or just closed
    #    - ≥ 6 months old, no winner → closed (award notice likely published separately)
    pub_date = flat.get("_pub_date")  # already a date object from _flatten_notice
    if isinstance(pub_date, date):
        days_old = (date.today() - pub_date).days
        return "Open" if 0 <= days_old < 180 else "Closed"

    # 5. National portal ID year fallback (e.g. CZ-N006/24/V00015605 → year 2024)
    tid = str(notice.get("tender_id", ""))
    m = re.search(r"[/_-](\d{2})[/_]", tid)
    if m:
        year = 2000 + int(m.group(1))
        if year <= date.today().year - 1:
            return "Closed"
        return "Open"

    return "Unknown"


_NATIONAL_PREFIXES = ("UK-", "NO-", "CZ-", "PL-", "DE-", "SE-", "FI-", "CA-")


def determine_source(notice: dict) -> str:
    """Return the correct Source label for a notice.

    Rule: TED notices keep "TED" (optionally suffixed with national portal code
    if they were enriched from a national source). National-only notices use
    their own source code.
    """
    tid = notice.get("tender_id", "")
    current = notice.get("source", "") or "TED"

    is_national_only = any(tid.startswith(p) for p in _NATIONAL_PREFIXES)
    if is_national_only:
        return current  # already correct (e.g. "CZ-NEN", "UK-CF")

    # TED notice — keep "TED", append national suffix if enriched from portal
    if current.startswith("TED+"):
        return current  # already tagged (e.g. "TED+NO-DF")
    if notice.get("source_url_national") and current not in ("TED", ""):
        return f"TED+{current}"
    return "TED"


# Column definitions: (header, field_key, width, data_type)
# data_type: "str", "date", "num", "int", "url"
# 23 columns B–X (Source + Source URL National added at W/X)
COLUMNS = [
    ("Tender ID",             "tender_id",               16, "str"),    # B
    ("Title",                 "_title_final",             55, "str"),    # C
    ("Country",               "_country_normalized",      14, "str"),    # D
    ("Authority",             "_authority_name",          35, "str"),    # E
    ("Publication Date",      "_pub_date",                16, "date"),   # F
    ("Status",                "_status",                  12, "str"),    # G
    ("Est. Value",            "_value_num",               18, "num"),    # H
    ("Currency",              "_value_currency",          10, "str"),    # I
    ("Est. Value (EUR)",      "_value_eur_num",           18, "num"),    # J
    # Slot 1 (K-M)
    ("Trailer Type (1)",      "_trailer_type_1_final",    35, "str"),    # K
    ("Category (1)",          "_trailer_cat_1_final",     18, "str"),    # L
    ("Quantity (1)",          "_trailer_qty_1_int",       14, "int"),    # M
    # Slot 2 (N-P)
    ("Trailer Type (2)",      "_trailer_type_2_final",    35, "str"),    # N
    ("Category (2)",          "_trailer_cat_2_final",     18, "str"),    # O
    ("Quantity (2)",          "_trailer_qty_2_int",       14, "int"),    # P
    # Supplemental
    ("Additional Equip.",     "_additional_equip_final",  35, "str"),    # Q
    ("Additional Qty",        "_additional_qty_int",      14, "int"),    # R
    ("Contract Duration",     "_contract_duration_final", 16, "str"),    # S
    ("Winner",                "_winner_name",             30, "str"),    # T
    # Source / URL columns
    ("Source URL (TED)",      "ted_url",                  45, "url"),    # U (renamed from "TED URL")
    ("Description",           "_description_final",       65, "str"),    # V
    ("Source",                "_source",                  14, "str"),    # W  ← NEW
    ("Source URL (National)", "_source_url_national",     45, "url"),    # X  ← NEW
]

# Header fill colors by column index (0-based within COLUMNS list)
_COL_FILLS = {
    12: "2E75B6",  # N – Slot 2 (medium blue)
    13: "2E75B6",
    14: "2E75B6",
    21: "548235",  # W – Source (green, visually distinct)
    22: "548235",  # X – Source URL (National)
}


class ExcelExporter:
    """Exports AI-classified notices with native Excel data types."""

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    DATA_FONT = Font(name="Aptos Narrow", size=11)
    LINK_FONT = Font(name="Aptos Narrow", size=11, color="0563C1", underline="single")
    TITLE_FONT = Font(name="Aptos Narrow", size=11, bold=True)

    ROW_HEIGHT = 45
    HEADER_ROW = 4
    DATA_START_ROW = 5
    COL_OFFSET = 2  # starts at column B
    SHEET_NAME = "Scraper Data"

    def __init__(self, config: dict):
        self.config = config
        self.base_dir = Path(config.get("output", {}).get("export_dir", "data/export"))
        self.archive_dir = self.base_dir / "archive"
        self.test_dir = self.base_dir / "test"
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.archive_dir.mkdir(exist_ok=True)
        self.test_dir.mkdir(exist_ok=True)
        self.vorlage_path = Path(__file__).parent.parent / "Vorlage.xlsx"

    def _flatten_notice(self, notice: dict) -> dict:
        """Flatten notice into row dict with proper data types.

        Handles both TED-format notices (title/description/contracting_authority)
        and national portal notices (DE-SB, PL-BZP) that use _title_final /
        _national_raw_text / _authority_name etc. as their primary fields.
        """
        flat = dict(notice)

        # Authority — national notices store name in _authority_name
        auth = notice.get("contracting_authority", {}) or {}
        flat["_authority_name"] = (auth.get("name_short") or auth.get("name", "")
                                   or notice.get("_authority_name", ""))

        # Country — national notices store in _country_normalized
        flat["_country_normalized"] = normalize_country(
            auth.get("country", "") or notice.get("_country_normalized", ""))

        # Title: AI English > _title_final (national) > raw TED title
        if notice.get("_title_english"):
            flat["_title_final"] = notice["_title_english"]
        elif notice.get("_title_final"):
            flat["_title_final"] = str(notice["_title_final"])
        else:
            title = notice.get("title", "")
            if isinstance(title, dict):
                title = title.get("eng") or title.get("deu") or next(iter(title.values()), "")
            flat["_title_final"] = str(title)

        # Publication date — national notices use _pub_date_clean
        pub_raw = str(notice.get("publication_date", "")
                      or notice.get("_pub_date_clean", "") or "")
        flat["_pub_date"] = parse_date(pub_raw)

        # Value — national notices store in _value_amount / _value_currency
        val = notice.get("estimated_value") or {}
        amount = clean_value(val.get("amount")) or clean_value(notice.get("_value_amount"))
        currency = val.get("currency", "") or notice.get("_value_currency", "")
        flat["_value_num"] = amount
        flat["_value_currency"] = currency

        # EUR conversion → float
        if amount is not None:
            rate = FX_RATES_TO_EUR.get(currency)
            if rate:
                flat["_value_eur_num"] = round(amount * rate)
            else:
                flat["_value_eur_num"] = None
        else:
            flat["_value_eur_num"] = None

        # Winner — national notices store in _winner_name directly
        award = notice.get("award") or {}
        flat["_winner_name"] = clean_winner(
            award.get("winner_name", "") or notice.get("_winner_name", ""))

        # AI-classified slot fields (slots 1 and 2 only)
        for slot in (1, 2):
            flat[f"_trailer_type_{slot}_final"] = notice.get(f"_trailer_type_{slot}_ai") or ""
            flat[f"_trailer_cat_{slot}_final"] = notice.get(f"_trailer_category_{slot}_ai") or ""
            # Fall back to legacy _trailer_quantity_ai (slot 1 only) if slot field is empty
            qty = notice.get(f"_trailer_quantity_{slot}_ai")
            if qty is None and slot == 1:
                qty = notice.get("_trailer_quantity_ai")
            flat[f"_trailer_qty_{slot}_int"] = clean_int(qty)

        flat["_contract_duration_final"] = notice.get("_contract_duration_ai", "")
        flat["_additional_equip_final"] = notice.get("_additional_equipment_ai", "")
        flat["_additional_qty_int"] = clean_int(notice.get("_additional_qty_ai"))

        # Description: AI English > _description_final (national) > raw TED desc
        if notice.get("_description_english"):
            flat["_description_final"] = notice["_description_english"][:500]
        elif notice.get("_description_final"):
            flat["_description_final"] = str(notice["_description_final"])[:500]
        else:
            desc = notice.get("description", "")
            if isinstance(desc, dict):
                desc = desc.get("eng") or desc.get("deu") or next(iter(desc.values()), "")
            flat["_description_final"] = str(desc or "")[:500]

        # Tender status (Open / Awarded / Closed / Unknown)
        flat["_status"] = determine_status(notice, flat)

        # Source tagging — use determine_source() for consistent labelling
        flat["_source"] = determine_source(notice)
        flat["_source_url_national"] = notice.get("source_url_national") or ""

        return flat

    def _generate_filename(self, test_mode: bool = False) -> Path:
        date_str = datetime.now().strftime("%y%m%d")
        target_dir = self.test_dir if test_mode else self.base_dir
        version = 1
        while True:
            fname = f"{date_str}_TED_Tender Data_00.{version:02d}.xlsx"
            if not (target_dir / fname).exists():
                return target_dir / fname
            version += 1

    def _archive_previous(self, new_path: Path):
        for old in self.base_dir.glob("*_TED_Tender Data_*.xlsx"):
            if old != new_path and old.is_file():
                try:
                    shutil.move(str(old), str(self.archive_dir / old.name))
                    logger.info(f"Archived: {old.name}")
                except Exception as e:
                    logger.warning(f"Could not archive {old.name}: {e}")

    def export(self, filtered_dir: str = "data/filtered",
               filename: Optional[str] = None,
               test_mode: bool = False,
               canada_notices: Optional[list] = None) -> str:
        """Export AI-classified notices to Excel with native data types."""
        filtered_path = Path(filtered_dir)
        relevant = self._load_json(filtered_path / "relevant.json")
        if not relevant:
            logger.warning("No relevant notices to export!")
            return ""
        relevant = self._dedup_for_export(relevant)

        # Output path
        if filename:
            output_path = (self.test_dir if test_mode else self.base_dir) / filename
        else:
            output_path = self._generate_filename(test_mode=test_mode)

        if not test_mode:
            self._archive_previous(output_path)

        # Copy Vorlage
        if self.vorlage_path.exists():
            for attempt in range(5):
                try:
                    shutil.copy2(self.vorlage_path, output_path)
                    break
                except PermissionError:
                    output_path = output_path.with_name(f"{output_path.stem}_v{attempt+2}.xlsx")
            wb = load_workbook(output_path)
            if hasattr(wb, '_external_links'):
                wb._external_links = []
            # Remove named ranges that cause Excel repair warnings
            try:
                for name in list(wb.defined_names):
                    del wb.defined_names[name]
            except Exception:
                pass
        else:
            wb = Workbook()
            wb.active.title = self.SHEET_NAME

        # Get sheet
        if self.SHEET_NAME in wb.sheetnames:
            ws = wb[self.SHEET_NAME]
        elif "Sheet3" in wb.sheetnames:
            ws = wb["Sheet3"]
            ws.title = self.SHEET_NAME
        else:
            ws = wb.create_sheet(self.SHEET_NAME)

        # Clear data rows
        for row in range(self.DATA_START_ROW, max(ws.max_row + 1, self.DATA_START_ROW + 1)):
            for col in range(1, 25):
                ws.cell(row=row, column=col).value = None

        # Title
        ws.cell(row=2, column=2, value="BPW Defense | Tender Portals").font = self.TITLE_FONT

        # Headers + column widths
        ws.column_dimensions["A"].width = 3
        for col_idx, (header, _, width, _) in enumerate(COLUMNS):
            col = self.COL_OFFSET + col_idx
            cell = ws.cell(row=self.HEADER_ROW, column=col, value=header)
            hex_color = _COL_FILLS.get(col_idx, "1F4E79")
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[self.HEADER_ROW].height = 35

        # ── Load unified blacklist ──────────────────────────────────────────────
        # Sources: config/blacklist.json (false positives, known duplicates, UK
        # training/sports) and legacy config/uk_blacklist.json.
        blacklist: set = set()
        _config_root = Path(__file__).parent.parent / "config"

        # New unified blacklist
        _bl_path = _config_root / "blacklist.json"
        if _bl_path.exists():
            try:
                _bl_data = json.loads(_bl_path.read_text(encoding="utf-8"))
                for _section in _bl_data.values():
                    if isinstance(_section, dict) and "ids" in _section:
                        blacklist.update(_section["ids"])
            except Exception as _e:
                logger.warning(f"Could not load blacklist.json: {_e}")

        # Legacy UK blacklist (for backward-compat; usually merged into blacklist.json)
        _uk_path = _config_root / "uk_blacklist.json"
        if _uk_path.exists():
            try:
                _uk_data = json.loads(_uk_path.read_text(encoding="utf-8"))
                blacklist.update(_uk_data.get("blacklisted_ids", []))
            except Exception as _e:
                logger.warning(f"Could not load uk_blacklist.json: {_e}")

        if blacklist:
            logger.info(f"Blacklist loaded: {len(blacklist)} IDs total")

        # ── Load manual category/field overrides ────────────────────────────────
        # config/manual_overrides.json: {tender_id: {field: value, ...}}
        # Applied after AI classification — permanent Opus-review corrections.
        manual_overrides: dict = {}
        _ov_path = _config_root / "manual_overrides.json"
        if _ov_path.exists():
            try:
                manual_overrides = json.loads(_ov_path.read_text(encoding="utf-8"))
                if manual_overrides:
                    logger.info(f"Manual overrides loaded: {len(manual_overrides)} entries")
            except Exception as _e:
                logger.warning(f"Could not load manual_overrides.json: {_e}")

        # Data rows
        thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))
        row_count = 0
        blacklisted_count = 0

        for notice in relevant:
            # Skip blacklisted IDs (false positives, duplicates, UK irrelevant)
            if notice.get("tender_id") in blacklist:
                blacklisted_count += 1
                continue

            # Apply manual overrides before flattening
            if manual_overrides and notice.get("tender_id") in manual_overrides:
                ov = manual_overrides[notice["tender_id"]]
                for _field, _val in ov.items():
                    if not _field.startswith("_reason"):
                        notice = {**notice, _field: _val}

            flat = self._flatten_notice(notice)
            if blacklisted_count and row_count == 0:
                logger.info(f"Export: {blacklisted_count} notices excluded by blacklist")

            # Skip if no trailer type in slot 1
            trailer_type = flat.get("_trailer_type_1_final", "")
            if not trailer_type or not str(trailer_type).strip():
                continue

            excel_row = self.DATA_START_ROW + row_count

            for col_idx, (_, field, _, dtype) in enumerate(COLUMNS):
                col = self.COL_OFFSET + col_idx
                value = flat.get(field)
                cell = ws.cell(row=excel_row, column=col)

                # Write with correct data type
                if dtype == "url" and value and str(value).startswith("http"):
                    cell.hyperlink = str(value)
                    cell.value = str(value)
                    cell.font = self.LINK_FONT
                elif dtype == "date":
                    if isinstance(value, date):
                        cell.value = value
                        cell.number_format = "YYYY-MM-DD"
                        cell.font = self.DATA_FONT
                    else:
                        cell.value = str(value) if value else ""
                        cell.font = self.DATA_FONT
                elif dtype == "num":
                    if isinstance(value, (int, float)) and value is not None:
                        cell.value = value
                        cell.number_format = '#,##0'
                        cell.font = self.DATA_FONT
                    else:
                        cell.value = None
                        cell.font = self.DATA_FONT
                elif dtype == "int":
                    if isinstance(value, int) and value is not None:
                        cell.value = value
                        cell.number_format = '#,##0'
                        cell.font = self.DATA_FONT
                    else:
                        cell.value = None
                        cell.font = self.DATA_FONT
                else:
                    cell.value = str(value) if value and value != "null" and value != "None" else ""
                    cell.font = self.DATA_FONT

                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

            ws.row_dimensions[excel_row].height = self.ROW_HEIGHT
            row_count += 1

        # Freeze panes disabled (was B5) — no freeze
        ws.freeze_panes = None

        # Auto-filter
        last_col = get_column_letter(self.COL_OFFSET + len(COLUMNS) - 1)
        ws.auto_filter.ref = f"B{self.HEADER_ROW}:{last_col}{self.DATA_START_ROW + row_count}"

        # ── Canada (Historical) sheet ──
        canada_row_count = 0
        if canada_notices:
            # Ensure sheet exists
            if self.CANADA_SHEET_NAME not in wb.sheetnames:
                wb.create_sheet(self.CANADA_SHEET_NAME)
            canada_row_count = self.export_canada_sheet(wb, canada_notices)

        wb.save(output_path)

        logger.info(f"Excel exported: {output_path}")
        logger.info(f"  Scraper Data rows: {row_count}")
        if canada_notices:
            logger.info(f"  Canada (Historical) rows: {canada_row_count}")

        # Always publish a fixed-name copy for GitHub / customer access.
        # Skipped for test exports so test runs don't overwrite the public file.
        if not test_mode:
            try:
                latest_path = self.base_dir / "TED_Defence_Trailers_LATEST.xlsx"
                shutil.copy2(output_path, latest_path)
                logger.info(f"Latest copy: {latest_path}")
            except Exception as e:
                logger.warning(f"Could not write LATEST copy: {e}")

        return str(output_path)

    def _dedup_for_export(self, notices: list) -> list:
        """Deduplicate by tender_id, keeping the notice with the most data."""
        seen: dict = {}
        for n in notices:
            tid = n.get("tender_id", "")
            if tid not in seen or self._data_score(n) > self._data_score(seen[tid]):
                seen[tid] = n

        result = list(seen.values())
        removed = len(notices) - len(result)
        if removed > 0:
            logger.info(f"Export dedup: {len(notices)} -> {len(result)} (removed {removed} duplicates by tender_id)")
        return result

    @staticmethod
    def _data_score(n: dict) -> int:
        score = 0
        if (n.get("award") or {}).get("winner_name"):
            score += 100
        val = (n.get("estimated_value") or {}).get("amount")
        try:
            if val and float(str(val).replace(",", "") or 0) > 0.01:
                score += 50
        except (ValueError, TypeError):
            pass
        if n.get("_trailer_quantity_1_ai"):
            score += 25
        score += len(str(n.get("publication_date", "")))
        return score

    # ── Canada (Historical) Sheet ──

    CANADA_SHEET_NAME = "Canada (Historical)"
    CANADA_COLUMNS = [
        ("Contract ID",       "tender_id",           20, "str"),   # B
        ("Title",             "title",               55, "str"),   # C
        ("Country",           "country",             12, "str"),   # D
        ("Authority",         "authority",           35, "str"),   # E
        ("Contract Date",     "date",                16, "date"),  # F
        ("Status",            "status",              12, "str"),   # G
        ("Value (CAD)",       "value_cad",           18, "num"),   # H
        ("Currency",          "currency",            10, "str"),   # I
        ("Value (EUR)",       "value_eur",           18, "num"),   # J
        ("Trailer Type (1)",  "trailer_type_1",      35, "str"),   # K
        ("Category (1)",      "trailer_category_1",  18, "str"),   # L
        ("Quantity (1)",      "trailer_quantity_1",  14, "int"),   # M
        ("Winner",            "winner",              30, "str"),   # N
        ("Source",            "source",              10, "str"),   # O
        ("Description",       "description",         65, "str"),   # P
    ]

    def export_canada_sheet(self, wb, canada_notices: list) -> int:
        """Write Canadian historical DND contracts to 'Canada (Historical)' sheet."""
        sheet_name = self.CANADA_SHEET_NAME

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # Clear existing data rows
        for row in range(5, max(ws.max_row + 1, 6)):
            for col in range(1, 18):
                ws.cell(row=row, column=col).value = None

        # Title
        ws.cell(row=2, column=2,
                value="BPW Defense | Canada DND Contracts (Historical)").font = self.TITLE_FONT

        # Headers
        ws.column_dimensions["A"].width = 3
        for col_idx, (header, _, width, _) in enumerate(self.CANADA_COLUMNS):
            col = self.COL_OFFSET + col_idx
            cell = ws.cell(row=self.HEADER_ROW, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[self.HEADER_ROW].height = 35

        thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))
        row_count = 0

        for notice in canada_notices:
            excel_row = self.DATA_START_ROW + row_count
            for col_idx, (_, field, _, dtype) in enumerate(self.CANADA_COLUMNS):
                col = self.COL_OFFSET + col_idx
                value = notice.get(field)
                cell = ws.cell(row=excel_row, column=col)

                if dtype == "date":
                    d = parse_date(str(value)) if value else None
                    if isinstance(d, date):
                        cell.value = d
                        cell.number_format = "YYYY-MM-DD"
                    else:
                        cell.value = str(value) if value else ""
                    cell.font = self.DATA_FONT
                elif dtype == "num":
                    v = clean_value(value)
                    cell.value = v
                    if v is not None:
                        cell.number_format = '#,##0'
                    cell.font = self.DATA_FONT
                elif dtype == "int":
                    v = clean_int(value)
                    cell.value = v
                    if v is not None:
                        cell.number_format = '#,##0'
                    cell.font = self.DATA_FONT
                else:
                    cell.value = str(value) if value and value not in ("null", "None") else ""
                    cell.font = self.DATA_FONT

                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

            ws.row_dimensions[excel_row].height = self.ROW_HEIGHT
            row_count += 1

        # Auto-filter
        if row_count > 0:
            last_col = get_column_letter(self.COL_OFFSET + len(self.CANADA_COLUMNS) - 1)
            ws.auto_filter.ref = (
                f"B{self.HEADER_ROW}:{last_col}{self.DATA_START_ROW + row_count}"
            )

        logger.info(f"Canada sheet written: {row_count} rows")
        return row_count

    @staticmethod
    def _load_json(path: Path) -> list:
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, list) else []
        return []
