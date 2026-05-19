"""
Quality Review Module - Opus-based post-run QA.

Reads the exported Excel, asks Claude Opus to flag:
  - Duplicates (same procurement exported twice under different tender IDs)
  - False positives (not actually trailers / not actually defence)
  - Category errors (wrong Trailer Category given the Type text)
  - Blacklist buzzwords (generic placeholders leaking into the Type field)
  - Extraction opportunities (rows where Additional Equipment / Qty
    clearly encode a second trailer type that should have been slot 2)

Results are saved as JSON to data/quality_review.json.

Trigger via: python main.py --all --review   OR   python main.py --uk --review
"""

from __future__ import annotations

import json
import logging
import os
from pathlib import Path
from typing import Optional

import requests
import urllib3

_SSL_VERIFY = os.environ.get("SSL_VERIFY_DISABLE", "").strip().lower() not in ("1", "true", "yes")
if not _SSL_VERIFY:
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logger = logging.getLogger(__name__)


class QualityReviewer:
    """Post-run quality reviewer backed by Claude Opus via OpenRouter."""

    API_URL = "https://openrouter.ai/api/v1/chat/completions"
    MODEL = "anthropic/claude-opus-4.6"

    def __init__(self):
        self.api_key = (
            os.environ.get("LLM_OPENROUTER_API_KEY")
            or os.environ.get("OPENROUTER_API_KEY")
            or None
        )
        if not self.api_key:
            logger.warning("LLM_OPENROUTER_API_KEY not set -- quality review disabled.")
        self.session = requests.Session()
        self.session.verify = _SSL_VERIFY
        self.session.headers.update({
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key or ''}",
            "HTTP-Referer": "https://bpw-tender-radar.internal",
            "X-Title": "BPW Defence Tender Radar",
        })

    @property
    def is_available(self) -> bool:
        return self.api_key is not None

    # ------------------------------------------------------------------ #
    # Excel -> row dicts                                                 #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _load_rows(excel_path: Path) -> list[dict]:
        """Return list of row dicts from the Excel export's data sheet."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl not installed -- cannot load Excel.")
            return []

        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active  # exporter writes to the active sheet
        # Header row is row 4 in the Vorlage template
        header_row = 4
        headers: list[str] = []
        for cell in ws[header_row]:
            headers.append(str(cell.value) if cell.value is not None else "")

        rows: list[dict] = []
        for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(v is None or v == "" for v in r):
                continue
            row = {}
            for h, v in zip(headers, r):
                if h:
                    row[h] = v
            if row.get("Tender ID"):
                rows.append(row)
        return rows

    # ------------------------------------------------------------------ #
    # Opus review call                                                   #
    # ------------------------------------------------------------------ #

    def _build_prompt(self, rows: list[dict]) -> str:
        """Compact the rows for the model and build the review prompt."""
        compact = []
        for r in rows:
            compact.append({
                "tender_id": r.get("Tender ID"),
                "title": r.get("Title"),
                "country": r.get("Country"),
                "authority": r.get("Authority"),
                "status": r.get("Status"),
                "value_eur": r.get("Est. Value (EUR)"),
                "trailer_type_1": r.get("Trailer Type (1)"),
                "category_1": r.get("Category (1)"),
                "qty_1": r.get("Quantity (1)"),
                "trailer_type_2": r.get("Trailer Type (2)"),
                "category_2": r.get("Category (2)"),
                "qty_2": r.get("Quantity (2)"),
                "additional": r.get("Additional Equip."),
                "additional_qty": r.get("Additional Qty"),
                "winner": r.get("Winner"),
                "source": r.get("Source"),
                "description": (r.get("Description") or "")[:400],
            })

        rows_json = json.dumps(compact, ensure_ascii=False)

        valid_categories = [
            "Low-Bed", "Semitrailer", "Dolly", "Tank Trailer",
            "Mission Module", "Loading System", "Special Purpose",
            "Ammunition Trailer", "Field Kitchen", "Cargo Trailer", "Other",
        ]

        return f"""You are a senior defence-procurement QA analyst. Review the following exported rows from the TED/UK defence-trailer scraper and flag issues.

VALID CATEGORIES: {json.dumps(valid_categories)}

Return STRICT JSON (no markdown, no prose) with this exact shape:
{{
  "duplicates": [
    {{"tender_ids": ["...", "..."], "reason": "same authority+title+year, both describe the same procurement"}}
  ],
  "false_positives": [
    {{"tender_id": "...", "reason": "why this is not a trailer / not defence"}}
  ],
  "category_errors": [
    {{"tender_id": "...", "current": "Cargo Trailer", "should_be": "Tank Trailer", "reason": "type text clearly says fuel tanker"}}
  ],
  "blacklist_buzzwords": [
    {{"tender_id": "...", "field": "trailer_type_1", "value": "Trailer", "reason": "generic placeholder"}}
  ],
  "extraction_opportunities": [
    {{"tender_id": "...", "reason": "Additional Equipment 'low-bed 40t' should be trailer_type_2 (slot 2)"}}
  ],
  "summary": {{
    "total_rows": {len(rows)},
    "issues_found": 0,
    "notes": "short human-readable summary"
  }}
}}

Only include entries that are actual issues. Empty arrays are fine. Be conservative — don't flag on weak evidence.

ROWS:
{rows_json}
"""

    def _call_opus(self, prompt: str) -> Optional[dict]:
        if not self.api_key:
            return None
        payload = {
            "model": self.MODEL,
            "max_tokens": 8000,
            "messages": [{"role": "user", "content": prompt}],
        }
        try:
            resp = self.session.post(self.API_URL, json=payload, timeout=180)
            if resp.status_code != 200:
                logger.error(f"Opus review HTTP {resp.status_code}: {resp.text[:300]}")
                return None
            data = resp.json()
            text = (data.get("choices") or [{}])[0].get("message", {}).get("content", "")
            text = text.strip()
            if text.startswith("```"):
                # Strip code fences
                text = text.split("\n", 1)[1] if "\n" in text else text
                if text.endswith("```"):
                    text = text.rsplit("```", 1)[0]
            return json.loads(text)
        except Exception as e:
            logger.error(f"Opus review failed: {e}")
            return None

    # ------------------------------------------------------------------ #
    # Public API                                                         #
    # ------------------------------------------------------------------ #

    def review(self, excel_path: Path, out_path: Optional[Path] = None) -> Optional[dict]:
        """Run the quality review and save JSON result."""
        excel_path = Path(excel_path)
        if not excel_path.exists():
            logger.error(f"Excel not found: {excel_path}")
            return None
        if not self.is_available:
            logger.warning("Quality review skipped: no LLM_OPENROUTER_API_KEY.")
            return None

        rows = self._load_rows(excel_path)
        logger.info(f"Quality review: loaded {len(rows)} rows from {excel_path.name}")
        if not rows:
            return {"summary": {"total_rows": 0, "issues_found": 0, "notes": "No rows to review"}}

        prompt = self._build_prompt(rows)
        result = self._call_opus(prompt)
        if result is None:
            return None

        if out_path is None:
            out_path = excel_path.parent.parent / "quality_review.json"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        logger.info(f"Quality review saved: {out_path}")
        return result
